import os
import openpyxl
import zipfile
import shutil
import glob
import xml.etree.ElementTree as ET
import numpy as np
from typing import Dict, Generator, List
from scipy.spatial import distance
from uuid import uuid4
from openpyxl.utils import coordinate_to_tuple
from PIL import Image
from langchain.docstore.document import Document
from langchain.document_loaders import (
    PyPDFLoader,
    UnstructuredWordDocumentLoader,
    WikipediaLoader,
)
from langchain.text_splitter import RecursiveCharacterTextSplitter
from dotenv import load_dotenv

load_dotenv()

text_splitter_chunk_size = int(os.environ.get("TEXT_SPLITTER_CHUNK_SIZE", 300))
text_splitter_chunk_overlap = int(os.environ.get("TEXT_SPLITTER_CHUNK_OVERLAP", 150))

text_splitter = RecursiveCharacterTextSplitter(
    chunk_size=text_splitter_chunk_size,
    chunk_overlap=text_splitter_chunk_overlap,
)

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "x": "http://schemas.openxmlformats.org/package/2006/relationships",
}

EMB_KEY = "{%s}embed" % NS["r"]
ANCHOR_LIST = ["xdr:oneCellAnchor", "xdr:twoCellAnchor", "xdr:absoluteAnchor"]
IMG_BASE_PATH = "img"


def load_pdf(path: str, uuid: str) -> List[Document]:
    documents = PyPDFLoader(path).load()
    for doc in documents:
        doc.metadata["doc_uuid"] = uuid
        doc.metadata["source"] = path.split("/")[-1]

    documents = text_splitter.split_documents(documents)
    for doc in documents:
        doc.page_content = "passage: " + doc.page_content

    return documents


def load_docx(path: str, uuid: str) -> List[Document]:
    documents = UnstructuredWordDocumentLoader(path).load()
    for doc in documents:
        print(doc.metadata)
        doc.metadata["source"] = path.split("/")[-1]
        doc.metadata["doc_uuid"] = uuid

    documents = text_splitter.split_documents(documents)
    for doc in documents:
        doc.page_content = "passage: " + doc.page_content

    return documents


def load_wiki(
    search_query: str, lang="ja", doc_content_chars_max=10000
) -> List[Document]:
    documents = WikipediaLoader(
        search_query, lang, doc_content_chars_max=doc_content_chars_max
    ).load()
    for doc in documents:
        doc.metadata["source"] = doc.metadata["title"]

    documents = text_splitter.split_documents(documents)
    for doc in documents:
        doc.page_content = (
            "passage: " + doc.metadata["title"] + " > " + doc.page_content
        )
    return documents


def _get_img_infos(file_path: str, tmp_path: str) -> List[Dict]:
    """
    Excel内の画像を保存し、画像のパスとセルの番地を返す
    Args:
        file_path (str): Excelファイルへのパス
        tmp_path (str): 作業フォルダ
    Returns:
        List[Dict]: 画像のパスとセルの番地
    """

    drawing_path = f"{tmp_path}/xl/drawings"

    # Excelファイルを一時ディレクトリに解凍
    with zipfile.ZipFile(file_path, "r") as zip_ref:
        zip_ref.extractall(tmp_path)

    drawings = glob.glob(f"{drawing_path}/*.xml")
    shape_infos = []
    for i, d in enumerate(drawings):
        for autoshape in _enumerate_autoshapes(d):
            # 画像の座標を取得
            from_col = int(autoshape.find("./xdr:from/xdr:col", NS).text)
            from_row = int(autoshape.find("./xdr:from/xdr:row", NS).text)
            shape_coord = np.asarray([from_row, from_col]) + 1

            # 画像の情報を取得してリストに追加
            for emb in autoshape.findall(".//a:blip", NS):
                rId = emb.attrib[EMB_KEY]

                img_rel_path = _get_rel_img(tmp_path, i + 1, rId)
                img_name = os.path.basename(img_rel_path)
                img_ext = img_name.split(".")[-1]

                new_img_name = f"{str(uuid4())}.{img_ext}"
                new_img_path = os.path.join(IMG_BASE_PATH, new_img_name)

                transform = autoshape.find("./xdr:pic/xdr:spPr/a:xfrm", NS)
                ext = transform.find("./a:ext", NS)
                r = int(transform.attrib["rot"]) if "rot" in transform.attrib else 0
                r /= 60000

                cx = int(int(ext.attrib["cx"]) / 10000)
                cy = int(int(ext.attrib["cy"]) / 10000)

                img_path = os.path.join(drawing_path, img_rel_path)

                img = Image.open(img_path).resize((cx, cy)).rotate(r, expand=True)
                img.save(new_img_path)
                img.close()

                shape_infos.append(
                    {"type": "image", "img_path": new_img_path, "coord": shape_coord}
                )

                print(img_path, shape_coord, cx, cy)

            # テキストの情報を取得してリストに追加
            texts = []
            for tx_body in autoshape.findall(".//xdr:txBody", NS):
                paragraphs = [para for para in tx_body.findall(".//a:p", NS)]
                for paragraph in paragraphs:
                    text = "".join(
                        [text.text for text in paragraph.findall(".//a:t", NS)]
                    )
                    texts.append(text)

                shape_infos.append(
                    {"type": "text", "text": "\n".join(texts), "coord": shape_coord}
                )
                print("\n".join(texts), shape_coord)
    return shape_infos


def load_xlsx(file_path: str, uuid: str) -> List[Document]:
    # 画像を保存するディレクトリがなければ作成
    if not os.path.exists("img"):
        os.mkdir("img")

    # 作業ディレクトリの設定
    work_dir = "tmp/work"
    tmp_path = os.path.join(work_dir, file_path)

    # 画像の情報を取得
    shape_infos = _get_img_infos(file_path, tmp_path)

    # Excelファイルを開く
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    documents = []

    # 各シートに対して処理
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # テキストの情報を取得
        cell_values = [
            {"value": x["text"], "coord": x["coord"]}
            for x in filter(lambda x: x["type"] == "text", shape_infos)
        ]

        # セルの値とその座標を取得
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None or len(str(cell.value)) == 0:
                    continue
                coord = np.asarray(coordinate_to_tuple(cell.coordinate))
                cell_values.append({"value": str(cell.value), "coord": coord})
                print(cell.value, coord)

        for cell_value in cell_values:
            cell_value["i"] = cell_value["coord"].tolist()

        # 座標でソート
        cell_values = sorted(cell_values, key=lambda x: x["i"])

        coords = [cell_value["coord"] for cell_value in cell_values]

        # 最も近いノードを取得
        closest_nodes = [
            _get_closest_node_index(node["coord"], coords) for node in shape_infos
        ]

        lines_with_img_path = []
        metadata = {
            "source": file_path.split("/")[-1],
            "sheet_name": sheet_name,
            "doc_uuid": uuid,
        }

        for i, cell_value in enumerate(cell_values):
            for r in filter(lambda x: x == i, closest_nodes):
                index = closest_nodes.index(r)
                shape_info = shape_infos[index]
                if shape_info["type"] == "image":
                    img_path = shape_info["img_path"]
                    metadata["img_path"] = img_path
                    cell_value["value"] += f"\n![img_name]({img_path})"
            lines_with_img_path.append(cell_value["value"])

        # ドキュメントを作成
        page_content = "\n".join(lines_with_img_path)
        documents.append(
            Document(
                page_content=f"{metadata['source']} > {page_content}", metadata=metadata
            )
        )

    # 作業ディレクトリの削除
    shutil.rmtree(work_dir)

    return documents


LOAD_DOC_FN = {"pdf": load_pdf, "docx": load_docx, "xlsx": load_xlsx}


def load_documents(file_path: str, uuid: str) -> List[Document]:
    ext = file_path.split(".")[-1]
    if not ext in LOAD_DOC_FN:
        raise ValueError("Document type must be in 'xlsx', 'docx', or 'pdf'")

    loader = LOAD_DOC_FN[ext]
    return loader(file_path, uuid)


def _get_rel_img(file_path: str, i: int, rId: str) -> str:
    rel_path = f"{file_path}/xl/drawings/_rels/drawing{i}.xml.rels"
    print(i, rel_path, rId)
    tree = ET.parse(rel_path)
    root = tree.getroot()

    # 画像の関連ファイルのパスを取得
    for rel in root.findall("./x:Relationship", NS):
        if rel.attrib["Id"] == rId:
            return rel.attrib["Target"]


def _enumerate_autoshapes(xml_path: str):
    """
    オートシェイプの情報を列挙する
    Args:
        xml_path (str): drawing情報を含むXMLファイルへのパス
    Yields:
        Element: オートシェイプ情報
    """

    tree = ET.parse(xml_path)
    root = tree.getroot()
    for anchor in ANCHOR_LIST:
        for autoshape in root.findall(anchor, NS):
            yield autoshape


def _get_closest_node_index(node, nodes) -> int:
    # 与えられたノードに最も近いノードのインデックスを返す
    closest_index = distance.cdist([node], nodes).argmin()
    return closest_index


def main():
    file_path = "TEST.xlsx"  # replace with your file path
    print(load_xlsx(file_path, str(uuid4())))


if __name__ == "__main__":
    main()
